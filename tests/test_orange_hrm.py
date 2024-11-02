import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from pages.orange_hrm_page import OrangeHRMPage
import openpyxl
from datetime import datetime


def read_test_data(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_data.append(row)
    workbook.close()
    return test_data


def write_test_result(filename, row, result):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet.cell(row=row + 2, column=7).value = result  # Column 7 is for Test Result
    # Fill in Date and Time of Test
    sheet.cell(row=row + 2, column=4).value = datetime.now().date()  # Date
    sheet.cell(row=row + 2, column=5).value = datetime.now().time()  # Time
    workbook.save(filename)
    workbook.close()


@pytest.mark.parametrize("test_id, username, password", read_test_data('data/login_data.xlsx'))
def test_login(test_id, username, password):
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    page = OrangeHRMPage(driver)

    try:
        driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

        page.enter_username(username)
        page.enter_password(password)
        page.click_login()

        # Check if login was successful
        if page.get_current_url() == "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index":
            result = "Pass"
        else:
            result = "Fail"

    except Exception as e:
        result = "Error: " + str(e)

    finally:
        write_test_result('data/login_data.xlsx', test_id, result)
        driver.quit()
